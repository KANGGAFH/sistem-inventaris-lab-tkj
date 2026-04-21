# ============================================================
# app.py - Lab Inventory Management System v3.0
# Sekolah : SMKN 1 Cikarang Selatan, Bekasi - Lab TJKT
# Fitur   : CRUD, QR Code, Export Excel/PDF, Cetak Label QR,
#           Filter & Sorting, Peminjaman & Pengembalian Barang
# ============================================================

from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file)
from flask_mysqldb import MySQL
import bcrypt, qrcode, os, json, io
from datetime import datetime, date, timedelta
from functools import wraps

# ─────────────────────────────────────────────
# INISIALISASI APP
# ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'labkomputer_secret_kp_2026'
app.config['MYSQL_HOST']        = 'localhost'
app.config['MYSQL_USER']        = 'root'
app.config['MYSQL_PASSWORD']    = ''
app.config['MYSQL_DB']          = 'inventaris_lab'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'
mysql = MySQL(app)

# ─────────────────────────────────────────────
# CONTEXT PROCESSOR
# ─────────────────────────────────────────────
@app.context_processor
def inject_globals():
    return {'now_year': datetime.now().strftime('%Y')}

# ============================================================
# REFERENSI INVENTARIS TETAP (Non-Consumable)
# Format kode: [PREFIX_KAT]-[TAHUN]-[NOURUT]
# Tipe produk diisi MANUAL oleh pengguna (tidak ada dropdown)
# Contoh: KOM-2024-0001, NET-2024-0003, MON-2024-0002
# ============================================================
KATEGORI_PREFIX = {
    'Komputer'  : 'KOM',
    'Monitor'   : 'MON',
    'Jaringan'  : 'NET',
    'Peripheral': 'PRP',
    'Printer'   : 'PRN',
    'Listrik'   : 'PWR',
    'Scanner'   : 'SCN',
    'Furnitur'  : 'FRN',
    'Lainnya'   : 'LNY',
}

JENIS_ASET = ['Aset Tetap', 'Aset Bergerak']

# ============================================================
# REFERENSI BARANG HABIS PAKAI (Consumable)
# Format kode: CSM-[PREFIX_KAT]-[TAHUN]-[NOURUT]
# Contoh: CSM-KBL-2024-0001 (Kabel & Konektor)
# ============================================================
KATEGORI_CONSUMABLE = {
    'Kabel & Konektor' : 'KBL',
    'Alat & Komponen'  : 'ALT',
    'ATK & Kertas'     : 'KRT',
    'Bahan Praktek'    : 'BHN',
    'Lainnya'          : 'LNY',
}

SATUAN_LIST = ['pcs', 'meter', 'roll', 'pack', 'box', 'rim', 'set', 'unit', 'lembar', 'botol']

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Silakan login terlebih dahulu.','warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def generate_kode_barang(kategori):
    """Format: [PREFIX_KAT]-[TAHUN]-[NOURUT]  contoh: KOM-2024-0001
    
    Menggunakan MAX nomor urut tertinggi yang pernah ada (bukan COUNT),
    sehingga penghapusan barang tidak menyebabkan duplikasi kode.
    """
    prefix = KATEGORI_PREFIX.get(kategori, 'LNY')
    tahun  = datetime.now().strftime('%Y')
    pola   = f'{prefix}-{tahun}-%'
    cur    = mysql.connection.cursor()
    # Ambil nomor urut tertinggi yang PERNAH ADA, bukan jumlah baris
    cur.execute(
        """SELECT COALESCE(
               MAX(CAST(SUBSTRING_INDEX(kode_barang, '-', -1) AS UNSIGNED)),
               0) AS max_n
           FROM barang
           WHERE kode_barang LIKE %s""",
        (pola,)
    )
    max_n = int(cur.fetchone()['max_n'])  # cast Decimal → int agar format :04d tidak error
    cur.close()
    # Pastikan kode baru belum terpakai (loop sampai dapat yang bebas)
    while True:
        max_n += 1
        kode_baru = f"{prefix}-{tahun}-{max_n:04d}"
        cur2 = mysql.connection.cursor()
        cur2.execute("SELECT id FROM barang WHERE kode_barang = %s", (kode_baru,))
        exists = cur2.fetchone(); cur2.close()
        if not exists:
            return kode_baru

def generate_kode_consumable(kategori):
    """Format: CSM-[PREFIX_KAT]-[TAHUN]-[NOURUT]  contoh: CSM-KBL-2024-0001
    
    Menggunakan MAX nomor urut tertinggi yang pernah ada,
    bukan COUNT, agar aman setelah penghapusan data.
    """
    prefix = KATEGORI_CONSUMABLE.get(kategori, 'LNY')
    tahun  = datetime.now().strftime('%Y')
    pola   = f'CSM-{prefix}-{tahun}-%'
    cur    = mysql.connection.cursor()
    cur.execute(
        """SELECT COALESCE(
               MAX(CAST(SUBSTRING_INDEX(kode_barang, '-', -1) AS UNSIGNED)),
               0) AS max_n
           FROM barang_habis_pakai
           WHERE kode_barang LIKE %s""",
        (pola,)
    )
    max_n = int(cur.fetchone()['max_n'])  # cast Decimal → int agar format :04d tidak error
    cur.close()
    while True:
        max_n += 1
        kode_baru = f"CSM-{prefix}-{tahun}-{max_n:04d}"
        cur2 = mysql.connection.cursor()
        cur2.execute("SELECT id FROM barang_habis_pakai WHERE kode_barang = %s", (kode_baru,))
        exists = cur2.fetchone(); cur2.close()
        if not exists:
            return kode_baru

def generate_kode_pinjam():
    today = datetime.now().strftime('%Y%m%d')
    cur   = mysql.connection.cursor()
    # Peminjaman juga pakai MAX agar aman
    cur.execute(
        """SELECT COALESCE(
               MAX(CAST(SUBSTRING_INDEX(kode_pinjam, '-', -1) AS UNSIGNED)),
               0) AS max_n
           FROM peminjaman
           WHERE kode_pinjam LIKE %s""",
        (f'PJM-{today}-%',))
    max_n = int(cur.fetchone()['max_n'])  # cast Decimal → int agar format :04d tidak error
    cur.close()
    while True:
        max_n += 1
        kode_baru = f"PJM-{today}-{max_n:04d}"
        cur2 = mysql.connection.cursor()
        cur2.execute("SELECT id FROM peminjaman WHERE kode_pinjam = %s", (kode_baru,))
        exists = cur2.fetchone(); cur2.close()
        if not exists:
            return kode_baru

def generate_kode_unit(kode_barang, nomor_unit):
    """Format: [KODE_BARANG]-U[NOURUT_2DIGIT]  contoh: TAB-2024-0001-U07"""
    n = int(nomor_unit)
    return f"{kode_barang}-U{n:02d}"

def buat_qr_unit(kode_unit):
    """QR Code untuk unit barang — link ke /unit/detail/<kode_unit>."""
    qr_dir = os.path.join(app.static_folder, 'qr')
    os.makedirs(qr_dir, exist_ok=True)
    url  = f"{_get_server_url()}/unit/detail/{kode_unit}"
    qr   = qrcode.QRCode(version=1,
                          error_correction=qrcode.constants.ERROR_CORRECT_H,
                          box_size=10, border=4)
    qr.add_data(url); qr.make(fit=True)
    img  = qr.make_image(fill_color="#1e3a8a", back_color="white")
    fn   = f"qr_{kode_unit}.png"
    img.save(os.path.join(qr_dir, fn))
    return f"qr/{fn}"

def _get_server_url():
    """Dapatkan base URL server secara otomatis dari request context."""
    try:
        from flask import request as _req
        host = _req.host  # contoh: 192.168.1.5:5000
        return f"http://{host}"
    except Exception:
        return "http://localhost:5000"

def buat_qr_code(kode_barang):
    """QR untuk inventaris tetap — link ke /barang/detail/<kode>."""
    qr_dir = os.path.join(app.static_folder, 'qr')
    os.makedirs(qr_dir, exist_ok=True)
    url  = f"{_get_server_url()}/barang/detail/{kode_barang}"
    qr   = qrcode.QRCode(version=1,
                          error_correction=qrcode.constants.ERROR_CORRECT_H,
                          box_size=10, border=4)
    qr.add_data(url); qr.make(fit=True)
    img  = qr.make_image(fill_color="#1a1a2e", back_color="white")
    fn   = f"qr_{kode_barang}.png"
    img.save(os.path.join(qr_dir, fn))
    return f"qr/{fn}"

def buat_qr_consumable(kode_barang):
    """QR untuk barang habis pakai — link ke /consumable/detail/<kode>."""
    qr_dir = os.path.join(app.static_folder, 'qr')
    os.makedirs(qr_dir, exist_ok=True)
    url  = f"{_get_server_url()}/consumable/detail/{kode_barang}"
    qr   = qrcode.QRCode(version=1,
                          error_correction=qrcode.constants.ERROR_CORRECT_H,
                          box_size=10, border=4)
    qr.add_data(url); qr.make(fit=True)
    img  = qr.make_image(fill_color="#065f46", back_color="white")
    fn   = f"qr_{kode_barang}.png"
    img.save(os.path.join(qr_dir, fn))
    return f"qr/{fn}"

# ─────────────────────────────────────────────────────────────
# HELPER: Filter waktu — digunakan di riwayat peminjaman & stok
# ─────────────────────────────────────────────────────────────
PERIOD_LABELS = {
    'today'  : 'Hari Ini',
    '7days'  : '7 Hari Terakhir',
    '1month' : '1 Bulan Terakhir',
    '1year'  : '1 Tahun Terakhir',
    ''       : 'Semua Waktu',
}

def get_date_range(period):
    """Kembalikan (date_from, label) berdasarkan kode period."""
    now = datetime.now()
    if period == 'today':
        return now.replace(hour=0, minute=0, second=0, microsecond=0), 'Hari Ini'
    elif period == '7days':
        return now - timedelta(days=7), '7 Hari Terakhir'
    elif period == '1month':
        return now - timedelta(days=30), '1 Bulan Terakhir'
    elif period == '1year':
        return now - timedelta(days=365), '1 Tahun Terakhir'
    return None, 'Semua Waktu'

def build_filter(search, kategori, jenis, kondisi, sort, order):
    clauses, params = [], []
    if search:
        clauses.append("(nama_barang LIKE %s OR kode_barang LIKE %s OR tipe LIKE %s)")
        like = f"%{search}%"; params += [like,like,like]
    if kategori: clauses.append("kategori=%s"); params.append(kategori)
    if jenis:    clauses.append("jenis_aset=%s"); params.append(jenis)
    if kondisi:  clauses.append("kondisi=%s"); params.append(kondisi)
    where = ("WHERE "+" AND ".join(clauses)) if clauses else ""
    safe_sort  = sort  if sort  in {'nama_barang','kode_barang','kategori','kondisi','jenis_aset','created_at'} else 'created_at'
    safe_order = 'ASC' if order=='asc' else 'DESC'
    return where, params, f"ORDER BY {safe_sort} {safe_order}"

# ============================================================
# AUTH
# ============================================================
@app.route('/', methods=['GET','POST'])
@app.route('/login', methods=['GET','POST'])
def login():
    if 'user_id' in session: return redirect(url_for('dashboard'))
    if request.method == 'POST':
        u = request.form.get('username','').strip()
        p = request.form.get('password','').strip()
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE username=%s",(u,))
        user = cur.fetchone(); cur.close()
        if user and bcrypt.checkpw(p.encode(), user['password'].encode()):
            session.update({'user_id':user['id'],'username':user['username'],
                            'nama_lengkap':user['nama_lengkap'],'role':user['role']})
            flash(f'Selamat datang, {user["nama_lengkap"]}!','success')
            return redirect(url_for('dashboard'))
        flash('Username atau password salah.','danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout.','info')
    return redirect(url_for('login'))

# ============================================================
# DASHBOARD
# ============================================================
@app.route('/dashboard')
@login_required
def dashboard():
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) AS n FROM barang"); total = cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang WHERE kondisi='Baik'"); baik=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang WHERE kondisi='Rusak Ringan'"); r_ringan=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang WHERE kondisi='Rusak Berat'"); r_berat=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM peminjaman WHERE status='Dipinjam'"); dipinjam=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM peminjaman WHERE status='Terlambat'"); terlambat=cur.fetchone()['n']
    # Consumable stats
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai"); csm_total=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai WHERE stok_sekarang <= stok_minimum AND stok_sekarang > 0"); csm_rendah=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai WHERE stok_sekarang = 0"); csm_habis=cur.fetchone()['n']
    cur.execute("SELECT * FROM barang ORDER BY created_at DESC LIMIT 5"); barang_terbaru=cur.fetchall()
    cur.execute("""SELECT p.*,b.nama_barang FROM peminjaman p
                   LEFT JOIN barang b ON p.barang_id=b.id
                   WHERE p.status='Dipinjam' ORDER BY p.tgl_kembali_rencana ASC LIMIT 5""")
    pinjam_aktif = cur.fetchall()
    cur.execute("SELECT kategori,COUNT(*) AS n FROM barang GROUP BY kategori ORDER BY n DESC")
    chart_kat = cur.fetchall()
    cur.execute("SELECT kondisi,COUNT(*) AS n FROM barang GROUP BY kondisi")
    chart_kondisi = cur.fetchall()
    cur.execute("SELECT jenis_aset,COUNT(*) AS n FROM barang GROUP BY jenis_aset")
    chart_aset = cur.fetchall()
    cur.close()

    # Auto-update status terlambat
    _update_terlambat()

    return render_template('dashboard.html',
        stats=dict(total=total,baik=baik,rusak_ringan=r_ringan,rusak_berat=r_berat,
                   dipinjam=dipinjam,terlambat=terlambat,
                   csm_total=csm_total,csm_rendah=csm_rendah,csm_habis=csm_habis),
        barang_terbaru=barang_terbaru,
        pinjam_aktif=pinjam_aktif,
        chart_kat=json.dumps([dict(r) for r in chart_kat]),
        chart_kondisi=json.dumps([dict(r) for r in chart_kondisi]),
        chart_aset=json.dumps([dict(r) for r in chart_aset]),
    )

def _update_terlambat():
    """Auto-update status jadi Terlambat jika sudah lewat waktu rencana kembali."""
    try:
        cur = mysql.connection.cursor()
        cur.execute("""UPDATE peminjaman SET status='Terlambat'
                       WHERE status='Dipinjam'
                       AND tgl_kembali_rencana < NOW()""")
        mysql.connection.commit(); cur.close()
    except: pass

# ============================================================
# BARANG — LIST
# ============================================================
@app.route('/barang')
@login_required
def barang_list():
    search   = request.args.get('search','').strip()
    f_kat    = request.args.get('kategori','').strip()
    f_jenis  = request.args.get('jenis_aset','').strip()
    f_kondisi= request.args.get('kondisi','').strip()
    sort     = request.args.get('sort','created_at')
    order    = request.args.get('order','desc')
    page     = int(request.args.get('page',1))
    per_page = 10; offset=(page-1)*per_page

    where,params,order_sql = build_filter(search,f_kat,f_jenis,f_kondisi,sort,order)
    cur = mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang {where} {order_sql} LIMIT %s OFFSET %s",
                params+[per_page,offset])
    rows = cur.fetchall()
    cur.execute(f"SELECT COUNT(*) AS n FROM barang {where}", params)
    total = cur.fetchone()['n']
    cur.close()

    return render_template('barang.html',
        barang_list=rows, search=search, f_kat=f_kat,
        f_jenis=f_jenis, f_kondisi=f_kondisi,
        sort=sort, order=order, next_order='asc' if order=='desc' else 'desc',
        page=page, total_pages=max(1,(total+per_page-1)//per_page),
        total_rows=total,
        kategori_list=list(KATEGORI_PREFIX.keys()),
        jenis_aset_list=JENIS_ASET,
    )

# ============================================================
# BARANG — TAMBAH
# ============================================================
@app.route('/barang/tambah', methods=['GET','POST'])
@login_required
def barang_tambah():
    if request.method=='POST':
        kategori   = request.form.get('kategori','Lainnya')
        tipe_input = request.form.get('tipe','').strip()
        kode       = generate_kode_barang(kategori)   # Format: KOM-2024-0001
        data = (kode,
                request.form['nama_barang'].strip(),
                kategori, tipe_input,
                request.form.get('jenis_aset','Aset Tetap'),
                request.form.get('spesifikasi','').strip(),
                request.form.get('kondisi','Baik'),
                int(request.form.get('jumlah',1)),
                request.form.get('tanggal_pengadaan') or None,
                buat_qr_code(kode),
                request.form.get('keterangan','').strip())
        cur = mysql.connection.cursor()
        cur.execute("""INSERT INTO barang
            (kode_barang,nama_barang,kategori,tipe,jenis_aset,
             spesifikasi,kondisi,jumlah,tanggal_pengadaan,qr_path,keterangan)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", data)
        mysql.connection.commit(); cur.close()
        flash(f'Barang berhasil ditambahkan dengan kode {kode}.','success')
        return redirect(url_for('barang_list'))
    return render_template('barang_form.html', mode='tambah', barang=None,
                           kategori_list=list(KATEGORI_PREFIX.keys()),
                           jenis_aset_list=JENIS_ASET)

# ============================================================
# BARANG — EDIT
# ============================================================
@app.route('/barang/edit/<int:id>', methods=['GET','POST'])
@login_required
def barang_edit(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE id=%s",(id,))
    barang = cur.fetchone()
    if not barang:
        flash('Barang tidak ditemukan.','danger')
        return redirect(url_for('barang_list'))
    if request.method=='POST':
        cur.execute("""UPDATE barang SET
            nama_barang=%s,kategori=%s,tipe=%s,jenis_aset=%s,
            spesifikasi=%s,kondisi=%s,jumlah=%s,
            tanggal_pengadaan=%s,keterangan=%s WHERE id=%s""",
            (request.form['nama_barang'].strip(),
             request.form.get('kategori'),
             request.form.get('tipe','').strip(),
             request.form.get('jenis_aset','Aset Tetap'),
             request.form.get('spesifikasi','').strip(),
             request.form.get('kondisi','Baik'),
             int(request.form.get('jumlah',1)),
             request.form.get('tanggal_pengadaan') or None,
             request.form.get('keterangan','').strip(),
             id))
        mysql.connection.commit(); cur.close()
        flash('Barang berhasil diperbarui.','success')
        return redirect(url_for('barang_list'))
    cur.close()
    return render_template('barang_form.html', mode='edit', barang=barang,
                           kategori_list=list(KATEGORI_PREFIX.keys()),
                           jenis_aset_list=JENIS_ASET)

# ============================================================
# BARANG — HAPUS
# ============================================================
@app.route('/barang/hapus/<int:id>', methods=['POST'])
@login_required
def barang_hapus(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT nama_barang,qr_path FROM barang WHERE id=%s",(id,))
    b = cur.fetchone()
    if b:
        # Simpan nama barang ke kolom nama_barang di semua tabel riwayat
        # sebelum barang dihapus, agar riwayat tetap terbaca walau barang sudah tidak ada
        cur.execute("""UPDATE peminjaman
                        SET barang_id = NULL
                        WHERE barang_id = %s""", (id,))
        cur.execute("""UPDATE riwayat_scan
                        SET barang_id = NULL
                        WHERE barang_id = %s""", (id,))
        # Hapus file QR
        if b['qr_path']:
            f = os.path.join(app.static_folder, b['qr_path'])
            if os.path.exists(f): os.remove(f)
        cur.execute("DELETE FROM barang WHERE id=%s",(id,))
        mysql.connection.commit()
        flash(f'Barang "{b["nama_barang"]}" berhasil dihapus. Riwayat pemakaian tetap tersimpan.','success')
    cur.close()
    return redirect(url_for('barang_list'))

# ============================================================
# BARANG — DETAIL (target QR scan)
# ============================================================
@app.route('/barang/detail/<kode>')
def barang_detail(kode):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE kode_barang=%s",(kode,))
    barang = cur.fetchone()
    if not barang: return render_template('404.html'),404

    # Catat riwayat scan
    cur.execute("""INSERT INTO riwayat_scan
        (barang_id,kode_barang,nama_barang,ip_scanner,user_agent)
        VALUES(%s,%s,%s,%s,%s)""",
        (barang['id'],kode,barang['nama_barang'],
         request.remote_addr,request.user_agent.string))
    mysql.connection.commit()

    # Cek apakah barang ini sedang dipinjam (untuk tampilkan tombol kembalikan)
    cur.execute("""SELECT p.id, p.kode_pinjam, p.nama_peminjam,
                          p.kelas_jabatan, p.tgl_pinjam, p.tgl_kembali_rencana,
                          p.status
                   FROM peminjaman p
                   WHERE p.barang_id=%s AND p.status IN ('Dipinjam','Terlambat')
                   ORDER BY p.tgl_pinjam DESC LIMIT 1""",
                (barang['id'],))
    peminjaman_aktif = cur.fetchone()

    cur.execute("SELECT * FROM riwayat_scan WHERE barang_id=%s ORDER BY scanned_at DESC LIMIT 10",
                (barang['id'],))
    riwayat = cur.fetchall(); cur.close()
    return render_template('detail.html', barang=barang, riwayat=riwayat,
                           peminjaman_aktif=peminjaman_aktif)

# ============================================================
# CETAK LABEL QR
# ============================================================
@app.route('/barang/cetak-label')
@login_required
def cetak_label():
    ids_raw = request.args.get('ids','')
    cur     = mysql.connection.cursor()
    if ids_raw:
        ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
        fmt = ','.join(['%s']*len(ids))
        cur.execute(f"SELECT * FROM barang WHERE id IN ({fmt}) AND qr_path IS NOT NULL", ids)
    else:
        cur.execute("SELECT * FROM barang WHERE qr_path IS NOT NULL ORDER BY kategori,nama_barang")
    rows = cur.fetchall(); cur.close()
    return render_template('cetak_label.html', barang_list=rows)

# ============================================================
# SCANNER QR
# ============================================================
@app.route('/scanner')
@login_required
def scanner():
    return render_template('scanner.html')

@app.route('/api/barang/<kode>')
def api_barang(kode):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE kode_barang=%s",(kode,))
    b = cur.fetchone()
    if b:
        for k in ('tanggal_pengadaan','created_at','updated_at'):
            if b.get(k): b[k]=str(b[k])
        # Sertakan data peminjaman aktif jika ada (hanya untuk barang tanpa unit tracking)
        pinjam = None
        if not b.get('has_unit_tracking'):
            cur.execute("""SELECT id, kode_pinjam, nama_peminjam,
                                  kelas_jabatan, tgl_pinjam, tgl_kembali_rencana, status
                           FROM peminjaman
                           WHERE barang_id=%s AND status IN ('Dipinjam','Terlambat')
                           ORDER BY tgl_pinjam DESC LIMIT 1""", (b['id'],))
            pinjam = cur.fetchone()
            if pinjam:
                for k in ('tgl_pinjam','tgl_kembali_rencana'):
                    if pinjam.get(k): pinjam[k] = pinjam[k].strftime('%d %b %Y, %H:%M')

        b['peminjaman_aktif'] = pinjam

        # Sertakan daftar unit jika has_unit_tracking
        if b.get('has_unit_tracking'):
            cur.execute("""SELECT id, kode_unit, nomor_unit, label_unit,
                                  kondisi, status
                           FROM barang_unit WHERE barang_id=%s
                           ORDER BY CAST(nomor_unit AS UNSIGNED)""", (b['id'],))
            units = cur.fetchall()
            b['units'] = units
        else:
            b['units'] = []

        cur.close()
        return jsonify({'status':'found','data':b})
    cur.close()
    return jsonify({'status':'not_found'}),404

# ============================================================
# PEMINJAMAN — LIST
# ============================================================
@app.route('/peminjaman')
@login_required
def peminjaman_list():
    _update_terlambat()
    f_status = request.args.get('status','').strip()
    search   = request.args.get('search','').strip()
    page     = int(request.args.get('page',1))
    per_page = 10; offset=(page-1)*per_page

    clauses,params=[],[]
    if search:
        clauses.append("(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s OR b.nama_barang LIKE %s)")
        like=f"%{search}%"; params+=[like,like,like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    where=("WHERE "+" AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT p.*,b.nama_barang,b.kategori,b.tipe,b.kode_barang AS kode_brg
                    FROM peminjaman p LEFT JOIN barang b ON p.barang_id=b.id
                    {where} ORDER BY p.created_at DESC LIMIT %s OFFSET %s""",
                params+[per_page,offset])
    rows = cur.fetchall()
    cur.execute(f"""SELECT COUNT(*) AS n FROM peminjaman p
                    LEFT JOIN barang b ON p.barang_id=b.id {where}""", params)
    total = cur.fetchone()['n']

    # Stats
    cur.execute("SELECT COUNT(*) AS n FROM peminjaman WHERE status='Dipinjam'");   s_dipinjam=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM peminjaman WHERE status='Dikembalikan'");s_kembali=cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM peminjaman WHERE status='Terlambat'");  s_terlambat=cur.fetchone()['n']
    cur.close()

    return render_template('peminjaman.html',
        pinjam_list=rows, search=search, f_status=f_status,
        page=page, total_pages=max(1,(total+per_page-1)//per_page),
        total_rows=total,
        stats=dict(dipinjam=s_dipinjam,kembali=s_kembali,terlambat=s_terlambat),
    )

# ============================================================
# PEMINJAMAN — TAMBAH
# ============================================================
@app.route('/peminjaman/tambah', methods=['GET','POST'])
@login_required
def peminjaman_tambah():
    cur = mysql.connection.cursor()
    # Hanya barang kondisi Baik yang bisa dipinjam
    cur.execute("SELECT id,kode_barang,nama_barang,kategori,tipe,jumlah,kondisi FROM barang WHERE kondisi IN ('Baik','Rusak Ringan') ORDER BY nama_barang")
    barang_tersedia = cur.fetchall()

    if request.method=='POST':
        barang_id  = int(request.form['barang_id'])
        jml_pinjam = int(request.form.get('jumlah_pinjam',1))

        # Cek stok
        cur.execute("SELECT kode_barang,nama_barang,jumlah FROM barang WHERE id=%s",(barang_id,))
        b = cur.fetchone()

        # Cek sudah berapa yang sedang dipinjam
        cur.execute("SELECT COALESCE(SUM(jumlah_pinjam),0) AS n FROM peminjaman WHERE barang_id=%s AND status IN ('Dipinjam','Terlambat')",(barang_id,))
        sdh_pinjam = cur.fetchone()['n']
        sisa = b['jumlah'] - sdh_pinjam

        if jml_pinjam > sisa:
            flash(f'Stok tidak cukup. Tersedia: {sisa} unit.','danger')
            cur.close()
            preselect_id = request.form.get('barang_id','')
            return render_template('peminjaman_form.html',
                                   barang_tersedia=barang_tersedia, mode='tambah', data=None,
                                   today=datetime.now().strftime('%Y-%m-%d'),
                                   preselect_id=preselect_id)

        kode_pinjam = generate_kode_pinjam()

        # Gabungkan tanggal + jam menjadi DATETIME string
        tgl_pinjam_dt          = f"{request.form['tgl_pinjam']} {request.form['jam_pinjam']}:00"
        tgl_kembali_rencana_dt = f"{request.form['tgl_kembali_rencana']} {request.form['jam_kembali_rencana']}:00"

        unit_id   = request.form.get('unit_id', '') or None
        kode_unit = None
        if unit_id:
            unit_id = int(unit_id)
            cur.execute("SELECT kode_unit FROM barang_unit WHERE id=%s",(unit_id,))
            u = cur.fetchone()
            if u: kode_unit = u['kode_unit']

        cur.execute("""INSERT INTO peminjaman
            (kode_pinjam,barang_id,kode_barang,nama_peminjam,kelas_jabatan,
             no_hp,keperluan,jumlah_pinjam,tgl_pinjam,tgl_kembali_rencana,
             unit_id,kode_unit,status)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Dipinjam')""",
            (kode_pinjam, barang_id, b['kode_barang'],
             request.form['nama_peminjam'].strip(),
             request.form.get('kelas_jabatan','').strip(),
             request.form.get('no_hp','').strip(),
             request.form.get('keperluan','').strip(),
             jml_pinjam,
             tgl_pinjam_dt, tgl_kembali_rencana_dt,
             unit_id, kode_unit))

        # Update status unit menjadi Dipinjam
        if unit_id:
            cur.execute("UPDATE barang_unit SET status='Dipinjam' WHERE id=%s",(unit_id,))

        mysql.connection.commit(); cur.close()
        flash(f'Peminjaman {kode_pinjam} berhasil dicatat.','success')
        return redirect(url_for('peminjaman_list'))

    # Ambil data unit untuk setiap barang yang has_unit_tracking
    units_per_barang = {}
    for b in barang_tersedia:
            cur2 = mysql.connection.cursor()
            cur2.execute("""SELECT id, kode_unit, nomor_unit, label_unit, kondisi
                             FROM barang_unit
                             WHERE barang_id=%s AND status='Tersedia'
                               AND kondisi != 'Rusak Berat'
                             ORDER BY CAST(nomor_unit AS UNSIGNED)""", (b['id'],))
            rows_unit = cur2.fetchall()
            cur2.close()
            # Konversi ke list of dict biasa agar tojson berjalan
            units_per_barang[str(b['id'])] = [
                {'id': u['id'], 'kode_unit': u['kode_unit'],
                 'nomor_unit': u['nomor_unit'], 'label_unit': u['label_unit'] or '',
                 'kondisi': u['kondisi']}
                for u in rows_unit
            ]
    cur.close()

    # Support preselect dari scan QR
    preselect_id   = request.args.get('barang_id', '')
    preselect_unit = request.args.get('unit_id', '')
    return render_template('peminjaman_form.html',
                           barang_tersedia=barang_tersedia, mode='tambah', data=None,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           preselect_id=preselect_id,
                           preselect_unit=preselect_unit,
                           units_per_barang=units_per_barang)

# ============================================================
# PEMINJAMAN — KEMBALIKAN
# ============================================================
@app.route('/peminjaman/kembalikan/<int:id>', methods=['GET','POST'])
@login_required
def peminjaman_kembalikan(id):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT p.*,b.nama_barang,b.tipe FROM peminjaman p
                   LEFT JOIN barang b ON p.barang_id=b.id WHERE p.id=%s""",(id,))
    pinjam = cur.fetchone()
    if not pinjam or pinjam['status']=='Dikembalikan':
        flash('Data tidak valid atau sudah dikembalikan.','warning')
        cur.close()
        return redirect(url_for('peminjaman_list'))

    if request.method=='POST':
        # Gabungkan tanggal + jam kembali aktual
        tgl_aktual_dt   = f"{request.form['tgl_kembali_aktual']} {request.form['jam_kembali_aktual']}:00"
        kondisi_kembali = request.form.get('kondisi_kembali','Baik')
        catatan         = request.form.get('catatan_admin','').strip()

        cur.execute("""UPDATE peminjaman SET
            status='Dikembalikan', tgl_kembali_aktual=%s,
            kondisi_kembali=%s, catatan_admin=%s
            WHERE id=%s""",
            (tgl_aktual_dt, kondisi_kembali, catatan, id))

        # Update kondisi barang jika kondisi kembali lebih buruk
        cur.execute("SELECT kondisi FROM barang WHERE id=%s",(pinjam['barang_id'],))
        kondisi_skrg = cur.fetchone()['kondisi']
        urutan = {'Baik':0,'Rusak Ringan':1,'Rusak Berat':2}
        if urutan.get(kondisi_kembali,0) > urutan.get(kondisi_skrg,0):
            cur.execute("UPDATE barang SET kondisi=%s WHERE id=%s",
                        (kondisi_kembali, pinjam['barang_id']))

        # Kembalikan status unit ke Tersedia jika barang pakai unit tracking
        if pinjam.get('unit_id'):
            cur.execute("""UPDATE barang_unit
                           SET status='Tersedia', kondisi=%s
                           WHERE id=%s""",
                        (kondisi_kembali, pinjam['unit_id']))

        mysql.connection.commit(); cur.close()
        flash(f'Barang "{pinjam["nama_barang"]}" berhasil dicatat kembali.','success')
        return redirect(url_for('peminjaman_list'))

    cur.close()
    today_now = datetime.now()
    return render_template('peminjaman_kembalikan.html', pinjam=pinjam,
                           today_date=today_now.strftime('%Y-%m-%d'),
                           today_time=today_now.strftime('%H:%M'))

# ============================================================
# PEMINJAMAN — DETAIL
# ============================================================
@app.route('/peminjaman/detail/<int:id>')
@login_required
def peminjaman_detail(id):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT p.*,b.nama_barang,b.kategori,b.tipe,b.kode_barang AS kode_brg,b.qr_path
                   FROM peminjaman p LEFT JOIN barang b ON p.barang_id=b.id
                   WHERE p.id=%s""",(id,))
    pinjam = cur.fetchone(); cur.close()
    if not pinjam:
        flash('Data peminjaman tidak ditemukan.','danger')
        return redirect(url_for('peminjaman_list'))
    today = date.today()
    return render_template('peminjaman_detail.html', pinjam=pinjam, today=today)

# ============================================================
# PEMINJAMAN — HAPUS
# ============================================================
@app.route('/peminjaman/hapus/<int:id>', methods=['POST'])
@login_required
def peminjaman_hapus(id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM peminjaman WHERE id=%s",(id,))
    mysql.connection.commit(); cur.close()
    flash('Data peminjaman berhasil dihapus.','success')
    return redirect(url_for('peminjaman_list'))

# api/tipe dihapus — tipe produk kini diisi manual oleh pengguna

# ============================================================
# EXPORT EXCEL
# ============================================================
@app.route('/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    search=request.args.get('search',''); f_kat=request.args.get('kategori','')
    f_jenis=request.args.get('jenis_aset',''); f_kondisi=request.args.get('kondisi','')
    where,params,order_sql = build_filter(search,f_kat,f_jenis,f_kondisi,'kategori','asc')
    cur = mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang {where} {order_sql}", params)
    rows = cur.fetchall(); cur.close()

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "Inventaris Lab TJKT"
    hfill = PatternFill("solid",fgColor="1e3a8a")
    hfont = Font(name='Calibri',bold=True,color="FFFFFF",size=11)
    center = Alignment(horizontal='center',vertical='center',wrap_text=True)
    left   = Alignment(horizontal='left',  vertical='center',wrap_text=True)
    thin   = Side(style='thin',color='CBD5E1')
    bdr    = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells('A1:K1'); ws['A1']='LAPORAN INVENTARIS LAB TJKT — SMKN 1 Cikarang Selatan'
    ws['A1'].font=Font(name='Calibri',bold=True,size=14,color="1e3a8a"); ws['A1'].alignment=center
    ws.merge_cells('A2:K2'); ws['A2']=f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    ws['A2'].font=Font(name='Calibri',size=10,italic=True,color="94a3b8"); ws['A2'].alignment=center
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=18

    headers=['No','Kode Barang','Nama Barang','Kategori','Tipe Produk','Jenis Aset','Spesifikasi','Kondisi','Jml','Tgl Pengadaan','Keterangan']
    widths  =[5, 20, 30, 14, 18, 14, 35, 14, 6, 16, 28]
    ws.append([]); ws.append(headers)
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(row=4,column=ci)
        cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=bdr
        ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[4].height=22

    kfill={'Baik':'D1FAE5','Rusak Ringan':'FEF3C7','Rusak Berat':'FEE2E2'}
    for i,b in enumerate(rows,1):
        tgl=b['tanggal_pengadaan'].strftime('%d/%m/%Y') if b.get('tanggal_pengadaan') else '-'
        ws.append([i,b['kode_barang'],b['nama_barang'],b['kategori'],
                   b['tipe'] or '-', b['jenis_aset'] or '-',
                   b['spesifikasi'] or '-', b['kondisi'],
                   b['jumlah'], tgl, b['keterangan'] or '-'])
        rn=4+i; ws.row_dimensions[rn].height=18
        for ci in range(1,12):
            cell=ws.cell(row=rn,column=ci); cell.border=bdr
            cell.alignment=center if ci in (1,4,5,6,8,9,10) else left
            if ci==8:
                cell.fill=PatternFill("solid",fgColor=kfill.get(b['kondisi'],'FFFFFF'))
                cell.font=Font(name='Calibri',bold=True,size=10)
    ws.freeze_panes='A5'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"Inventaris_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================
# EXPORT PDF
# ============================================================
@app.route('/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    search=request.args.get('search',''); f_kat=request.args.get('kategori','')
    f_jenis=request.args.get('jenis_aset',''); f_kondisi=request.args.get('kondisi','')
    where,params,order_sql = build_filter(search,f_kat,f_jenis,f_kondisi,'kategori','asc')
    cur=mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang {where} {order_sql}",params)
    rows=cur.fetchall(); cur.close()

    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),
                          leftMargin=1.5*cm,rightMargin=1.5*cm,
                          topMargin=2*cm,bottomMargin=1.5*cm)
    styles=getSampleStyleSheet(); elems=[]
    ts=ParagraphStyle('t',parent=styles['Title'],fontSize=14,
                      textColor=colors.HexColor('#1e3a8a'),spaceAfter=4)
    ss=ParagraphStyle('s',parent=styles['Normal'],fontSize=10,
                      textColor=colors.HexColor('#475569'),spaceAfter=2,alignment=1)
    ds=ParagraphStyle('d',parent=styles['Normal'],fontSize=8,
                      textColor=colors.HexColor('#94a3b8'),spaceAfter=12,alignment=1)
    elems+=[Paragraph('LAPORAN INVENTARIS LAB TJKT',ts),
            Paragraph('SMKN 1 Cikarang Selatan, Bekasi',ss),
            Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}',ds),
            Spacer(1,0.3*cm)]

    header=['No','Kode Barang','Nama Barang','Kategori','Tipe','Jenis Aset','Kondisi','Jml','Tgl Pengadaan']
    data=[header]
    for i,b in enumerate(rows,1):
        tgl=b['tanggal_pengadaan'].strftime('%d/%m/%Y') if b.get('tanggal_pengadaan') else '-'
        data.append([str(i),b['kode_barang'],b['nama_barang'],b['kategori'],
                     b['tipe'] or '-', b['jenis_aset'] or '-',
                     b['kondisi'],str(b['jumlah']),tgl])

    cw=[1*cm,3.5*cm,5.5*cm,3*cm,3.5*cm,3*cm,2.8*cm,1.2*cm,2.8*cm]
    tbl=Table(data,colWidths=cw,repeatRows=1)
    cfill={'Baik':colors.HexColor('#D1FAE5'),'Rusak Ringan':colors.HexColor('#FEF3C7'),
           'Rusak Berat':colors.HexColor('#FEE2E2')}
    ts_style=[('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e3a8a')),
              ('TEXTCOLOR',(0,0),(-1,0),colors.white),
              ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
              ('FONTSIZE',(0,0),(-1,0),9),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),
              ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('FONTSIZE',(0,1),(-1,-1),8),
              ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
              ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F8FAFF')]),
              ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#CBD5E1')),
              ('TOPPADDING',(0,0),(-1,-1),5),
              ('BOTTOMPADDING',(0,0),(-1,-1),5),
              ('ALIGN',(2,1),(2,-1),'LEFT')]
    for ri,b in enumerate(rows,1):
        fc=cfill.get(b['kondisi'])
        if fc: ts_style.append(('BACKGROUND',(6,ri),(6,ri),fc))
    tbl.setStyle(TableStyle(ts_style)); elems.append(tbl)
    doc.build(elems); buf.seek(0)
    fname=f"Inventaris_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf,as_attachment=True,download_name=fname,mimetype='application/pdf')

# ============================================================
# RIWAYAT SCAN
# ============================================================
@app.route('/riwayat')
@login_required
def riwayat_scan():
    cur=mysql.connection.cursor()
    cur.execute("""SELECT r.*,b.nama_barang,b.kategori FROM riwayat_scan r
                   LEFT JOIN barang b ON r.barang_id=b.id
                   ORDER BY r.scanned_at DESC LIMIT 100""")
    riwayat=cur.fetchall(); cur.close()
    return render_template('riwayat.html',riwayat=riwayat)

# ============================================================
# BARANG HABIS PAKAI (CONSUMABLE) — LIST
# ============================================================
@app.route('/consumable')
@login_required
def consumable_list():
    search   = request.args.get('search','').strip()
    f_kat    = request.args.get('kategori','').strip()
    f_stok   = request.args.get('stok','').strip()   # "rendah" = stok <= minimum
    page     = int(request.args.get('page',1))
    per_page = 10; offset = (page-1)*per_page

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append("(nama_barang LIKE %s OR kode_barang LIKE %s)")
        params += [like, like]
    if f_kat:
        clauses.append("kategori=%s"); params.append(f_kat)
    if f_stok == 'rendah':
        clauses.append("stok_sekarang <= stok_minimum")

    where    = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    cur      = mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang_habis_pakai {where} ORDER BY kategori,nama_barang LIMIT %s OFFSET %s",
                params + [per_page, offset])
    rows = cur.fetchall()
    cur.execute(f"SELECT COUNT(*) AS n FROM barang_habis_pakai {where}", params)
    total = cur.fetchone()['n']

    # Stats — stok_rendah TIDAK mencakup stok = 0 (habis punya card sendiri)
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai"); total_item = cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai WHERE stok_sekarang > 0 AND stok_sekarang <= stok_minimum"); stok_rendah = cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_habis_pakai WHERE stok_sekarang = 0"); habis = cur.fetchone()['n']
    cur.close()

    return render_template('consumable.html',
        rows=rows, search=search, f_kat=f_kat, f_stok=f_stok,
        page=page, total_pages=max(1,(total+per_page-1)//per_page),
        total_rows=total,
        stats=dict(total=total_item, rendah=stok_rendah, habis=habis),
        kategori_list=list(KATEGORI_CONSUMABLE.keys()),
        satuan_list=SATUAN_LIST,
    )

# ============================================================
# BARANG HABIS PAKAI — TAMBAH
# ============================================================
@app.route('/consumable/tambah', methods=['GET','POST'])
@login_required
def consumable_tambah():
    if request.method == 'POST':
        kategori    = request.form.get('kategori','Lainnya')
        kode        = generate_kode_consumable(kategori)
        nama        = request.form['nama_barang'].strip()
        satuan      = request.form.get('satuan','pcs')
        stok_awal   = int(request.form.get('stok_awal', 0))
        stok_min    = int(request.form.get('stok_minimum', 5))
        harga       = int(request.form.get('harga_satuan', 0) or 0)
        lokasi      = request.form.get('lokasi_simpan','').strip()
        ket         = request.form.get('keterangan','').strip()

        # Generate QR Code untuk consumable
        qr_path = buat_qr_consumable(kode)

        cur = mysql.connection.cursor()
        cur.execute("""INSERT INTO barang_habis_pakai
            (kode_barang,nama_barang,kategori,satuan,stok_awal,stok_sekarang,
             stok_minimum,harga_satuan,lokasi_simpan,qr_path,keterangan)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (kode,nama,kategori,satuan,stok_awal,stok_awal,stok_min,harga,lokasi,qr_path,ket))

        barang_id = cur.lastrowid
        if stok_awal > 0:
            cur.execute("""INSERT INTO riwayat_stok
                (barang_id,kode_barang,nama_barang,tipe_transaksi,jumlah,stok_sebelum,stok_sesudah,keterangan,dicatat_oleh)
                VALUES(%s,%s,%s,'Masuk',%s,0,%s,'Stok awal pengadaan',%s)""",
                (barang_id,kode,nama,stok_awal,stok_awal,session.get('nama_lengkap','Admin')))

        mysql.connection.commit(); cur.close()
        flash(f'Barang habis pakai "{nama}" berhasil ditambahkan dengan kode {kode}.','success')
        return redirect(url_for('consumable_list'))

    return render_template('consumable_form.html', mode='tambah', data=None,
                           kategori_list=list(KATEGORI_CONSUMABLE.keys()),
                           satuan_list=SATUAN_LIST,
                           now=datetime.now().strftime('%Y'))

# ============================================================
# BARANG HABIS PAKAI — EDIT
# ============================================================
@app.route('/consumable/edit/<int:id>', methods=['GET','POST'])
@login_required
def consumable_edit(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang_habis_pakai WHERE id=%s",(id,))
    data = cur.fetchone()
    if not data:
        flash('Data tidak ditemukan.','danger')
        return redirect(url_for('consumable_list'))
    if request.method == 'POST':
        cur.execute("""UPDATE barang_habis_pakai SET
            nama_barang=%s,kategori=%s,satuan=%s,stok_minimum=%s,
            harga_satuan=%s,lokasi_simpan=%s,keterangan=%s
            WHERE id=%s""",
            (request.form['nama_barang'].strip(),
             request.form.get('kategori','Lainnya'),
             request.form.get('satuan','pcs'),
             int(request.form.get('stok_minimum',5)),
             int(request.form.get('harga_satuan',0) or 0),
             request.form.get('lokasi_simpan','').strip(),
             request.form.get('keterangan','').strip(),
             id))
        mysql.connection.commit(); cur.close()
        flash('Data berhasil diperbarui.','success')
        return redirect(url_for('consumable_list'))
    cur.close()
    return render_template('consumable_form.html', mode='edit', data=data,
                           kategori_list=list(KATEGORI_CONSUMABLE.keys()),
                           satuan_list=SATUAN_LIST,
                           now=datetime.now().strftime('%Y'))

# ============================================================
# BARANG HABIS PAKAI — HAPUS
# ============================================================
@app.route('/consumable/hapus/<int:id>', methods=['POST'])
@login_required
def consumable_hapus(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT nama_barang, qr_path FROM barang_habis_pakai WHERE id=%s",(id,))
    data = cur.fetchone()
    if data:
        # Null-out barang_id di riwayat_stok agar riwayat tidak ikut terhapus
        cur.execute("""UPDATE riwayat_stok
                        SET barang_id = NULL
                        WHERE barang_id = %s""", (id,))
        # Hapus file QR jika ada
        if data.get('qr_path'):
            f = os.path.join(app.static_folder, data['qr_path'])
            if os.path.exists(f): os.remove(f)
        cur.execute("DELETE FROM barang_habis_pakai WHERE id=%s",(id,))
        mysql.connection.commit()
        flash(f'Barang "{data["nama_barang"]}" berhasil dihapus. Riwayat transaksi tetap tersimpan.','success')
    cur.close()
    return redirect(url_for('consumable_list'))

# ============================================================
# BARANG HABIS PAKAI — TRANSAKSI STOK (Pakai / Restock / Koreksi)
# ============================================================
@app.route('/consumable/transaksi/<int:id>', methods=['GET','POST'])
@login_required
def consumable_transaksi(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang_habis_pakai WHERE id=%s",(id,))
    data = cur.fetchone()
    if not data:
        flash('Data tidak ditemukan.','danger')
        return redirect(url_for('consumable_list'))

    if request.method == 'POST':
        tipe          = request.form['tipe_transaksi']
        jumlah        = int(request.form['jumlah'])
        nama_pemakai  = request.form.get('nama_pemakai','').strip()
        mata_pelajaran= request.form.get('mata_pelajaran','').strip()
        kelas         = request.form.get('kelas','').strip()
        ket           = request.form.get('keterangan','').strip()
        stok_lama     = data['stok_sekarang']

        if tipe == 'Masuk':
            stok_baru = stok_lama + jumlah
        elif tipe == 'Keluar':
            if jumlah > stok_lama:
                flash(f'Stok tidak cukup! Stok tersedia: {stok_lama} {data["satuan"]}.','danger')
                cur.close()
                return redirect(url_for('consumable_transaksi', id=id))
            stok_baru = stok_lama - jumlah
        else:  # Koreksi
            stok_baru = jumlah

        cur.execute("UPDATE barang_habis_pakai SET stok_sekarang=%s WHERE id=%s",
                    (stok_baru, id))
        cur.execute("""INSERT INTO riwayat_stok
            (barang_id,kode_barang,nama_barang,tipe_transaksi,jumlah,
             stok_sebelum,stok_sesudah,nama_pemakai,mata_pelajaran,kelas,
             keterangan,dicatat_oleh)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (id, data['kode_barang'], data['nama_barang'],
             tipe, jumlah, stok_lama, stok_baru,
             nama_pemakai or None, mata_pelajaran or None, kelas or None,
             ket or None, session.get('nama_lengkap','Admin')))
        mysql.connection.commit(); cur.close()

        label = {'Masuk':'Restock','Keluar':'Pemakaian','Koreksi':'Koreksi Stok'}.get(tipe, tipe)
        flash(f'{label} berhasil dicatat. Stok: {stok_lama} → {stok_baru} {data["satuan"]}.','success')
        return redirect(url_for('consumable_list'))

    # Riwayat transaksi barang ini
    cur.execute("""SELECT * FROM riwayat_stok WHERE barang_id=%s
                   ORDER BY created_at DESC LIMIT 20""", (id,))
    riwayat = cur.fetchall(); cur.close()
    return render_template('consumable_transaksi.html', data=data, riwayat=riwayat)

# ============================================================
# BARANG HABIS PAKAI — RIWAYAT STOK SEMUA
# ============================================================
@app.route('/consumable/riwayat')
@login_required
def consumable_riwayat():
    f_tipe   = request.args.get('tipe',   '').strip()
    f_period = request.args.get('period', '').strip()
    search   = request.args.get('search', '').strip()

    date_from, period_label = get_date_range(f_period)

    clauses, params = [], []
    if f_tipe:
        clauses.append("r.tipe_transaksi=%s"); params.append(f_tipe)
    if date_from:
        clauses.append("r.created_at >= %s"); params.append(date_from)
    if search:
        like = f"%{search}%"
        clauses.append(
            "(r.nama_barang LIKE %s OR r.kode_barang LIKE %s"
            " OR r.nama_pemakai LIKE %s OR r.kelas LIKE %s)"
        )
        params += [like, like, like, like]
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT r.*, b.satuan FROM riwayat_stok r
                   LEFT JOIN barang_habis_pakai b ON r.barang_id=b.id
                   {where} ORDER BY r.created_at DESC LIMIT 500""", params)
    riwayat = cur.fetchall()

    # Stats untuk periode aktif (inisialisasi default 0 agar tidak error)
    s_masuk = s_keluar = s_koreksi = 0
    sp = [date_from] if date_from else []
    sw = "WHERE created_at >= %s" if date_from else ""
    cur.execute(f"SELECT COUNT(*) AS n FROM riwayat_stok {sw}", sp)
    s_total = cur.fetchone()['n']
    for tipe in ['Masuk','Keluar','Koreksi']:
        # Bangun sw2 dengan benar: WHERE xxx AND yyy ATAU WHERE yyy
        if sw:
            sw2 = sw + f" AND tipe_transaksi='{tipe}'"
        else:
            sw2 = f"WHERE tipe_transaksi='{tipe}'"
        cur.execute(f"SELECT COUNT(*) AS n FROM riwayat_stok {sw2}", sp)
        val = cur.fetchone()['n']
        if tipe=='Masuk':    s_masuk   = val
        elif tipe=='Keluar': s_keluar  = val
        else:                s_koreksi = val
    cur.close()

    return render_template('consumable_riwayat.html',
        riwayat=riwayat, f_tipe=f_tipe,
        f_period=f_period, period_label=period_label,
        search=search,
        stats=dict(total=s_total, masuk=s_masuk,
                   keluar=s_keluar, koreksi=s_koreksi),
        PERIOD_LABELS=PERIOD_LABELS,
    )


# ============================================================
# CONSUMABLE — DETAIL (target QR scan, publik tanpa login)
# ============================================================
@app.route('/consumable/detail/<kode>')
def consumable_detail(kode):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang_habis_pakai WHERE kode_barang=%s", (kode,))
    data = cur.fetchone()
    if not data:
        return render_template('404.html'), 404

    # Riwayat 10 transaksi terakhir
    cur.execute("""SELECT * FROM riwayat_stok WHERE barang_id=%s
                   ORDER BY created_at DESC LIMIT 10""", (data['id'],))
    riwayat = cur.fetchall(); cur.close()
    return render_template('consumable_detail.html', data=data, riwayat=riwayat)

# ============================================================
# CONSUMABLE — API (untuk AJAX scanner)
# ============================================================
@app.route('/api/consumable/<kode>')
def api_consumable(kode):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang_habis_pakai WHERE kode_barang=%s", (kode,))
    b = cur.fetchone(); cur.close()
    if b:
        for k in ('created_at', 'updated_at'):
            if b.get(k): b[k] = str(b[k])
        if b.get('harga_satuan'): b['harga_satuan'] = float(b['harga_satuan'])
        return jsonify({'status': 'found', 'data': b})
    return jsonify({'status': 'not_found'}), 404

# ============================================================
# CONSUMABLE — EXPORT EXCEL
# ============================================================
@app.route('/consumable/export/excel')
@login_required
def consumable_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    f_kat  = request.args.get('kategori', '')
    f_stok = request.args.get('stok', '')

    clauses, params = [], []
    if f_kat:   clauses.append("kategori=%s"); params.append(f_kat)
    if f_stok == 'rendah': clauses.append("stok_sekarang <= stok_minimum")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang_habis_pakai {where} ORDER BY kategori,nama_barang", params)
    rows = cur.fetchall(); cur.close()

    wb  = openpyxl.Workbook(); ws = wb.active
    ws.title = "Barang Habis Pakai"

    hfill  = PatternFill("solid", fgColor="065f46")
    hfont  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='CBD5E1')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:J1')
    ws['A1'] = 'LAPORAN BARANG HABIS PAKAI — LAB TJKT SMKN 1 Cikarang Selatan'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color="065f46")
    ws['A1'].alignment = center

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="94a3b8")
    ws['A2'].alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    headers = ['No','Kode Barang','Nama Barang','Kategori','Satuan',
               'Stok Awal','Stok Sekarang','Stok Minimum','Status','Lokasi Simpan']
    widths  = [5, 20, 32, 18, 8, 10, 14, 12, 14, 22]

    ws.append([]); ws.append(headers)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 22

    sfill = {'Aman': 'D1FAE5', 'Stok Rendah': 'FEF3C7', 'Habis': 'FEE2E2'}
    for i, b in enumerate(rows, 1):
        if b['stok_sekarang'] == 0:       status = 'Habis'
        elif b['stok_sekarang'] <= b['stok_minimum']: status = 'Stok Rendah'
        else:                             status = 'Aman'

        ws.append([i, b['kode_barang'], b['nama_barang'], b['kategori'],
                   b['satuan'], b['stok_awal'], b['stok_sekarang'],
                   b['stok_minimum'], status, b['lokasi_simpan'] or '-'])
        rn = 4 + i; ws.row_dimensions[rn].height = 18
        for ci in range(1, 11):
            cell = ws.cell(row=rn, column=ci); cell.border = bdr
            cell.alignment = center if ci in (1,4,5,6,7,8,10) else left
            if ci == 9:
                cell.fill = PatternFill("solid", fgColor=sfill.get(status, 'FFFFFF'))
                cell.font = Font(name='Calibri', bold=True, size=10)
    ws.freeze_panes = 'A5'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"HabisPakai_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================
# CONSUMABLE — EXPORT PDF
# ============================================================
@app.route('/consumable/export/pdf')
@login_required
def consumable_export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    f_kat  = request.args.get('kategori', '')
    f_stok = request.args.get('stok', '')

    clauses, params = [], []
    if f_kat:   clauses.append("kategori=%s"); params.append(f_kat)
    if f_stok == 'rendah': clauses.append("stok_sekarang <= stok_minimum")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"SELECT * FROM barang_habis_pakai {where} ORDER BY kategori,nama_barang", params)
    rows = cur.fetchall(); cur.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet(); elems = []
    green  = colors.HexColor('#065f46')
    ts     = ParagraphStyle('t', parent=styles['Title'], fontSize=13, textColor=green, spaceAfter=4)
    ss     = ParagraphStyle('s', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#475569'), spaceAfter=2, alignment=1)
    ds     = ParagraphStyle('d', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'), spaceAfter=12, alignment=1)

    elems += [
        Paragraph('LAPORAN BARANG HABIS PAKAI — LAB TJKT', ts),
        Paragraph('SMKN 1 Cikarang Selatan, Bekasi', ss),
        Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}', ds),
        Spacer(1, 0.3*cm),
    ]

    header = ['No','Kode Barang','Nama Barang','Kategori','Satuan','Stok Skrg','Stok Min','Status','Lokasi']
    data   = [header]
    for i, b in enumerate(rows, 1):
        if b['stok_sekarang'] == 0:       status = 'Habis'
        elif b['stok_sekarang'] <= b['stok_minimum']: status = 'Rendah'
        else:                             status = 'Aman'
        data.append([str(i), b['kode_barang'], b['nama_barang'], b['kategori'],
                     b['satuan'], str(b['stok_sekarang']), str(b['stok_minimum']),
                     status, b['lokasi_simpan'] or '-'])

    cw  = [1*cm, 3.5*cm, 5.5*cm, 3.2*cm, 1.8*cm, 2*cm, 2*cm, 2.2*cm, 3.3*cm]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    sfill = {'Aman': colors.HexColor('#D1FAE5'), 'Rendah': colors.HexColor('#FEF3C7'), 'Habis': colors.HexColor('#FEE2E2')}
    ts_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#065f46')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 9),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE',   (0,1), (-1,-1), 8),
        ('FONTNAME',   (0,1), (-1,-1), 'Helvetica'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F0FDF4')]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('ALIGN',      (2,1), (2,-1), 'LEFT'),
    ]
    for ri, b in enumerate(rows, 1):
        if b['stok_sekarang'] == 0:       s = 'Habis'
        elif b['stok_sekarang'] <= b['stok_minimum']: s = 'Rendah'
        else:                             s = 'Aman'
        fc = sfill.get(s)
        if fc: ts_style.append(('BACKGROUND', (7, ri), (7, ri), fc))

    tbl.setStyle(TableStyle(ts_style)); elems.append(tbl)
    doc.build(elems); buf.seek(0)
    fname = f"HabisPakai_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')

# ============================================================
# ERROR HANDLER & MAIN
# ============================================================
@app.errorhandler(404)
def not_found(e): return render_template('404.html'),404

# ============================================================
# EXPORT RIWAYAT PEMINJAMAN — EXCEL
# ============================================================
@app.route('/peminjaman/export/excel')
@login_required
def peminjaman_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    f_status = request.args.get('status', '').strip()
    search   = request.args.get('search', '').strip()

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append("(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s OR b.nama_barang LIKE %s)")
        params += [like, like, like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT p.*, b.nama_barang, b.tipe, b.kode_barang AS kode_brg
                    FROM peminjaman p LEFT JOIN barang b ON p.barang_id=b.id
                    {where} ORDER BY p.created_at DESC""", params)
    rows = cur.fetchall(); cur.close()

    wb  = openpyxl.Workbook(); ws = wb.active
    ws.title = "Riwayat Peminjaman"

    hfill  = PatternFill("solid", fgColor="1e3a8a")
    hfont  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='CBD5E1')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Judul
    ws.merge_cells('A1:M1')
    ws['A1'] = 'RIWAYAT PEMINJAMAN & PENGEMBALIAN — LAB TJKT SMKN 1 Cikarang Selatan'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color="1e3a8a")
    ws['A1'].alignment = center

    ws.merge_cells('A2:M2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    if f_status: ws['A2'] = ws['A2'].value + f'  |  Filter: {f_status}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="94a3b8")
    ws['A2'].alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    headers = ['No','Kode Pinjam','Nama Barang','Tipe Barang','Nama Peminjam',
               'Kelas/Jabatan','No HP','Jumlah','Waktu Pinjam',
               'Rencana Kembali','Waktu Kembali Aktual','Kondisi Kembali','Status']
    widths  = [5, 20, 28, 18, 22, 14, 14, 7, 20, 20, 22, 14, 13]

    ws.append([]); ws.append(headers)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 22

    sfill = {
        'Dipinjam':     'DBEAFE',
        'Dikembalikan': 'D1FAE5',
        'Terlambat':    'FEE2E2',
    }

    def fmt_dt(val):
        if not val: return '-'
        try:    return val.strftime('%d/%m/%Y %H:%M')
        except: return str(val)

    for i, p in enumerate(rows, 1):
        ws.append([
            i,
            p['kode_pinjam'],
            p['nama_barang'] or '-',
            p['tipe'] or '-',
            p['nama_peminjam'],
            p['kelas_jabatan'] or '-',
            p['no_hp'] or '-',
            p['jumlah_pinjam'],
            fmt_dt(p['tgl_pinjam']),
            fmt_dt(p['tgl_kembali_rencana']),
            fmt_dt(p['tgl_kembali_aktual']),
            p['kondisi_kembali'] or '-',
            p['status'],
        ])
        rn = 4 + i
        ws.row_dimensions[rn].height = 18
        for ci in range(1, 14):
            cell = ws.cell(row=rn, column=ci)
            cell.border = bdr
            cell.alignment = center if ci in (1, 8) else left
        # Warna kolom Status
        status_cell = ws.cell(row=rn, column=13)
        fc = sfill.get(p['status'])
        if fc:
            status_cell.fill = PatternFill("solid", fgColor=fc)
            status_cell.font = Font(name='Calibri', bold=True, size=10)
            status_cell.alignment = center

    ws.freeze_panes = 'A5'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"Peminjaman_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# EXPORT RIWAYAT PEMINJAMAN — PDF
# ============================================================
@app.route('/peminjaman/export/pdf')
@login_required
def peminjaman_export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    f_status = request.args.get('status', '').strip()
    f_period = request.args.get('period', '').strip()
    search   = request.args.get('search', '').strip()

    date_from, period_label = get_date_range(f_period)

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append("(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s OR b.nama_barang LIKE %s)")
        params += [like, like, like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    if date_from:
        clauses.append("p.tgl_pinjam >= %s"); params.append(date_from)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT p.*,
                        COALESCE(b.nama_barang, p.kode_barang) AS nama_barang,
                        COALESCE(b.tipe, '') AS tipe
                    FROM peminjaman p LEFT JOIN barang b ON p.barang_id=b.id
                    {where} ORDER BY p.tgl_pinjam DESC""", params)
    rows = cur.fetchall(); cur.close()

    def fmt_dt(val):
        if not val: return '-'
        try: return val.strftime('%d/%m/%Y %H:%M')
        except: return str(val)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.8*cm, bottomMargin=1.2*cm)
    elems = []; navy = colors.HexColor('#1e3a8a')
    ts_h = ParagraphStyle('h', fontSize=12, textColor=navy, alignment=1,
                           fontName='Helvetica-Bold', spaceAfter=3)
    ts_s = ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#475569'),
                           alignment=1, spaceAfter=2)
    ts_d = ParagraphStyle('d', fontSize=7.5, textColor=colors.HexColor('#94a3b8'),
                           alignment=1, spaceAfter=8)

    judul = 'RIWAYAT PEMINJAMAN & PENGEMBALIAN — LAB TJKT'
    if f_status: judul += f' | {f_status}'
    if f_period: judul += f' | {period_label}'
    elems += [
        Paragraph(judul, ts_h),
        Paragraph('SMKN 1 Cikarang Selatan, Bekasi', ts_s),
        Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")} | Total: {len(rows)} data', ts_d),
        Spacer(1, 0.2*cm),
    ]

    header = ['No','Kode Pinjam','Nama Barang','Nama Peminjam',
              'Kls/Jabatan','Jml','Waktu Pinjam','Rencana Kembali',
              'Waktu Kembali','Kond. Kembali','Status']
    data   = [header]
    sfill  = {'Dipinjam': colors.HexColor('#DBEAFE'),
               'Dikembalikan': colors.HexColor('#D1FAE5'),
               'Terlambat': colors.HexColor('#FEE2E2')}
    ts_style = [
        ('BACKGROUND',    (0,0), (-1,0), navy),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE',      (0,1), (-1,-1), 7.5),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID',          (0,0), (-1,-1), 0.35, colors.HexColor('#CBD5E1')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ALIGN',         (2,1), (4,-1), 'LEFT'),
    ]
    for i, p in enumerate(rows, 1):
        data.append([
            str(i), p['kode_pinjam'],
            p.get('nama_barang') or p.get('kode_barang') or '-',
            p['nama_peminjam'],
            p['kelas_jabatan'] or '-',
            str(p['jumlah_pinjam']),
            fmt_dt(p['tgl_pinjam']),
            fmt_dt(p['tgl_kembali_rencana']),
            fmt_dt(p['tgl_kembali_aktual']),
            p['kondisi_kembali'] or '-',
            p['status'],
        ])
        fc = sfill.get(p['status'])
        if fc:
            ts_style.append(('BACKGROUND', (10, i), (10, i), fc))
            ts_style.append(('FONTNAME',   (10, i), (10, i), 'Helvetica-Bold'))

    cw  = [0.6*cm,2.6*cm,3.8*cm,3.2*cm,2*cm,0.8*cm,2.4*cm,2.4*cm,2.4*cm,2*cm,1.9*cm]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle(ts_style))
    elems.append(tbl)
    doc.build(elems); buf.seek(0)
    fname = f"Peminjaman_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


# ============================================================
# EXPORT RIWAYAT STOK HABIS PAKAI — EXCEL
# ============================================================
@app.route('/consumable/riwayat/export/excel')
@login_required
def consumable_riwayat_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    f_tipe = request.args.get('tipe', '').strip()   # Masuk/Keluar/Koreksi
    f_period = request.args.get('period', '').strip()

    date_from, period_label = get_date_range(f_period)

    clauses, params = [], []
    if f_tipe:
        clauses.append("r.tipe_transaksi=%s"); params.append(f_tipe)
    if date_from:
        clauses.append("r.created_at >= %s"); params.append(date_from)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT r.*, b.satuan, b.kategori AS kat_barang
                    FROM riwayat_stok r
                    LEFT JOIN barang_habis_pakai b ON r.barang_id=b.id
                    {where} ORDER BY r.created_at DESC""", params)
    rows = cur.fetchall(); cur.close()

    wb  = openpyxl.Workbook(); ws = wb.active
    ws.title = "Riwayat Stok Habis Pakai"

    hfill  = PatternFill("solid", fgColor="065f46")
    hfont  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='CBD5E1')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:L1')
    ws['A1'] = 'RIWAYAT PEMAKAIAN & RESTOCK BARANG HABIS PAKAI — LAB TJKT SMKN 1 Cikarang Selatan'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color="065f46")
    ws['A1'].alignment = center

    ws.merge_cells('A2:L2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    if f_tipe: ws['A2'] = ws['A2'].value + f'  |  Filter: {f_tipe}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="94a3b8")
    ws['A2'].alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    headers = ['No','Waktu','Kode Barang','Nama Barang','Kategori','Tipe Transaksi',
               'Jumlah','Satuan','Stok Sebelum','Stok Sesudah',
               'Nama Pemakai','Kelas','Mata Pelajaran','Keterangan','Dicatat Oleh']
    widths  = [5, 18, 20, 30, 16, 14, 8, 8, 12, 12, 22, 12, 20, 24, 18]

    ws.append([]); ws.append(headers)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 22

    tfill = {
        'Masuk':   'D1FAE5',
        'Keluar':  'FEE2E2',
        'Koreksi': 'E0E7FF',
    }

    for i, r in enumerate(rows, 1):
        waktu = r['created_at'].strftime('%d/%m/%Y %H:%M') if r.get('created_at') else '-'
        ws.append([
            i, waktu,
            r['kode_barang'], r['nama_barang'] or '-',
            r.get('kat_barang') or '-',
            r['tipe_transaksi'],
            r['jumlah'], r.get('satuan') or '-',
            r['stok_sebelum'], r['stok_sesudah'],
            r.get('nama_pemakai') or '-',
            r.get('kelas') or '-',
            r.get('mata_pelajaran') or '-',
            r.get('keterangan') or '-',
            r.get('dicatat_oleh') or '-',
        ])
        rn = 4 + i
        ws.row_dimensions[rn].height = 18
        for ci in range(1, 16):
            cell = ws.cell(row=rn, column=ci)
            cell.border = bdr
            cell.alignment = center if ci in (1, 7, 8, 9, 10) else left
        # Warna kolom Tipe Transaksi
        tipe_cell = ws.cell(row=rn, column=6)
        fc = tfill.get(r['tipe_transaksi'])
        if fc:
            tipe_cell.fill = PatternFill("solid", fgColor=fc)
            tipe_cell.font = Font(name='Calibri', bold=True, size=10)
            tipe_cell.alignment = center

    ws.freeze_panes = 'A5'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"RiwayatStok_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# EXPORT RIWAYAT STOK HABIS PAKAI — PDF
# ============================================================
@app.route('/consumable/riwayat/export/pdf')
@login_required
def consumable_riwayat_export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    f_tipe   = request.args.get('tipe',   '').strip()
    f_period = request.args.get('period', '').strip()

    date_from, period_label = get_date_range(f_period)

    clauses, params = [], []
    if f_tipe:
        clauses.append("r.tipe_transaksi=%s"); params.append(f_tipe)
    if date_from:
        clauses.append("r.created_at >= %s"); params.append(date_from)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""SELECT r.*, b.satuan
                    FROM riwayat_stok r
                    LEFT JOIN barang_habis_pakai b ON r.barang_id=b.id
                    {where} ORDER BY r.created_at DESC""", params)
    rows = cur.fetchall(); cur.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.8*cm, bottomMargin=1.2*cm)
    elems = []; green = colors.HexColor('#065f46')
    ts_h = ParagraphStyle('h', fontSize=12, textColor=green, alignment=1,
                           fontName='Helvetica-Bold', spaceAfter=3)
    ts_s = ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#475569'),
                           alignment=1, spaceAfter=2)
    ts_d = ParagraphStyle('d', fontSize=7.5, textColor=colors.HexColor('#94a3b8'),
                           alignment=1, spaceAfter=8)

    judul = 'RIWAYAT PEMAKAIAN & RESTOCK BARANG HABIS PAKAI — LAB TJKT'
    if f_tipe:   judul += f' | {f_tipe}'
    if f_period: judul += f' | {period_label}'
    elems += [
        Paragraph(judul, ts_h),
        Paragraph('SMKN 1 Cikarang Selatan, Bekasi', ts_s),
        Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")} | Total: {len(rows)} data', ts_d),
        Spacer(1, 0.2*cm),
    ]

    header = ['No','Waktu','Nama Barang','Tipe','Jml','Sat.',
              'Sblm','Ssdh','Nama Pemakai','Kelas','Mata Pelajaran','Keterangan']
    data   = [header]
    tfill  = {'Masuk':   colors.HexColor('#D1FAE5'),
               'Keluar':  colors.HexColor('#FEE2E2'),
               'Koreksi': colors.HexColor('#E0E7FF')}
    ts_style = [
        ('BACKGROUND',    (0,0), (-1,0), green),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE',      (0,1), (-1,-1), 7.5),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F0FDF4')]),
        ('GRID',          (0,0), (-1,-1), 0.35, colors.HexColor('#CBD5E1')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ALIGN',         (2,1), (2,-1), 'LEFT'),
        ('ALIGN',         (8,1), (11,-1), 'LEFT'),
    ]
    for i, r in enumerate(rows, 1):
        waktu = r['created_at'].strftime('%d/%m/%Y %H:%M') if r.get('created_at') else '-'
        data.append([
            str(i), waktu,
            r.get('nama_barang') or '-',
            r['tipe_transaksi'],
            str(r['jumlah']),
            r.get('satuan') or '-',
            str(r['stok_sebelum']),
            str(r['stok_sesudah']),
            r.get('nama_pemakai') or '-',
            r.get('kelas') or '-',
            r.get('mata_pelajaran') or '-',
            r.get('keterangan') or '-',
        ])
        fc = tfill.get(r['tipe_transaksi'])
        if fc:
            ts_style.append(('BACKGROUND', (3, i), (3, i), fc))
            ts_style.append(('FONTNAME',   (3, i), (3, i), 'Helvetica-Bold'))

    cw = [0.6*cm,2.2*cm,4*cm,1.8*cm,0.8*cm,0.8*cm,
          1.1*cm,1.1*cm,3*cm,1.8*cm,2.8*cm,2.8*cm]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle(ts_style))
    elems.append(tbl)
    doc.build(elems); buf.seek(0)
    fname = f"RiwayatStok_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


# ============================================================
# PENCARIAN GLOBAL
# ============================================================
@app.route('/search')
@login_required
def global_search():
    q = request.args.get('q', '').strip()
    if not q:
        return redirect(url_for('dashboard'))

    like = f"%{q}%"
    cur  = mysql.connection.cursor()

    # ── Inventaris Tetap ─────────────────────────────────────
    cur.execute("""
        SELECT id, kode_barang, nama_barang, kategori, tipe,
               kondisi, jumlah, qr_path, 'inventaris' AS sumber
        FROM barang
        WHERE nama_barang LIKE %s OR kode_barang LIKE %s
           OR tipe LIKE %s OR spesifikasi LIKE %s OR keterangan LIKE %s
        ORDER BY nama_barang LIMIT 20
    """, (like, like, like, like, like))
    hasil_barang = cur.fetchall()

    # ── Barang Habis Pakai ───────────────────────────────────
    cur.execute("""
        SELECT id, kode_barang, nama_barang, kategori, satuan,
               stok_sekarang, stok_minimum, qr_path, 'consumable' AS sumber
        FROM barang_habis_pakai
        WHERE nama_barang LIKE %s OR kode_barang LIKE %s
           OR kategori LIKE %s OR keterangan LIKE %s
        ORDER BY nama_barang LIMIT 20
    """, (like, like, like, like))
    hasil_consumable = cur.fetchall()

    # ── Peminjaman / Pemakaian ───────────────────────────────
    cur.execute("""
        SELECT p.id, p.kode_pinjam,
               COALESCE(b.nama_barang, p.kode_barang) AS nama_barang,
               p.nama_peminjam, p.kelas_jabatan, p.keperluan,
               p.tgl_pinjam, p.tgl_kembali_aktual, p.status,
               CASE WHEN p.barang_id IS NULL THEN 1 ELSE 0 END AS barang_dihapus
        FROM peminjaman p
        LEFT JOIN barang b ON p.barang_id = b.id
        WHERE p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s
           OR b.nama_barang   LIKE %s OR p.kode_barang LIKE %s
           OR p.kelas_jabatan LIKE %s OR p.keperluan   LIKE %s
        ORDER BY p.created_at DESC LIMIT 20
    """, (like, like, like, like, like, like))
    hasil_pinjam = cur.fetchall()

    # ── Riwayat Stok Consumable ──────────────────────────────
    cur.execute("""
        SELECT r.id, r.kode_barang, r.nama_barang,
               r.tipe_transaksi, r.jumlah, r.stok_sesudah,
               r.nama_pemakai, r.kelas, r.mata_pelajaran,
               r.created_at, b.id AS barang_id_real
        FROM riwayat_stok r
        LEFT JOIN barang_habis_pakai b ON r.barang_id = b.id
        WHERE r.nama_pemakai  LIKE %s OR r.kode_barang LIKE %s
           OR r.nama_barang   LIKE %s OR r.kelas       LIKE %s
           OR r.mata_pelajaran LIKE %s OR r.keterangan LIKE %s
        ORDER BY r.created_at DESC LIMIT 10
    """, (like, like, like, like, like, like))
    hasil_stok = cur.fetchall()

    cur.close()

    total = (len(hasil_barang) + len(hasil_consumable)
             + len(hasil_pinjam) + len(hasil_stok))

    return render_template('search.html',
        q=q, total=total,
        hasil_barang=hasil_barang,
        hasil_consumable=hasil_consumable,
        hasil_pinjam=hasil_pinjam,
        hasil_stok=hasil_stok,
    )

# ============================================================
# UNIT BARANG — LIST (per barang induk)
# ============================================================
@app.route('/barang/<int:barang_id>/unit')
@login_required
def unit_list(barang_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE id=%s", (barang_id,))
    barang = cur.fetchone()
    if not barang:
        flash('Barang tidak ditemukan.','danger')
        return redirect(url_for('barang_list'))

    cur.execute("""SELECT * FROM barang_unit
                   WHERE barang_id=%s ORDER BY CAST(nomor_unit AS UNSIGNED)""",
                (barang_id,))
    units = cur.fetchall()

    # Stats
    cur.execute("SELECT COUNT(*) AS n FROM barang_unit WHERE barang_id=%s",(barang_id,))
    total = cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_unit WHERE barang_id=%s AND status='Tersedia'",(barang_id,))
    tersedia = cur.fetchone()['n']
    cur.execute("SELECT COUNT(*) AS n FROM barang_unit WHERE barang_id=%s AND kondisi='Rusak Berat'",(barang_id,))
    rusak = cur.fetchone()['n']
    cur.close()

    return render_template('barang_unit.html',
        barang=barang, units=units,
        stats=dict(total=total, tersedia=tersedia,
                   dipinjam=total-tersedia, rusak=rusak),
    )

# ============================================================
# UNIT BARANG — TAMBAH
# ============================================================
@app.route('/barang/<int:barang_id>/unit/tambah', methods=['GET','POST'])
@login_required
def unit_tambah(barang_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE id=%s",(barang_id,))
    barang = cur.fetchone()
    if not barang:
        flash('Barang tidak ditemukan.','danger')
        return redirect(url_for('barang_list'))

    if request.method == 'POST':
        mode = request.form.get('mode','single')

        if mode == 'bulk':
            # Tambah banyak unit sekaligus (misal: 1 sampai 15)
            dari  = int(request.form.get('dari', 1))
            sampai= int(request.form.get('sampai', 1))
            kondisi   = request.form.get('kondisi', 'Baik')
            keterangan= request.form.get('keterangan','').strip()
            added = 0; skipped = []
            for n in range(dari, sampai + 1):
                kode_unit = generate_kode_unit(barang['kode_barang'], n)
                # Cek duplikat
                cur.execute("SELECT id FROM barang_unit WHERE kode_unit=%s",(kode_unit,))
                if cur.fetchone():
                    skipped.append(str(n)); continue
                qr = buat_qr_unit(kode_unit)
                cur.execute("""INSERT INTO barang_unit
                    (barang_id,kode_unit,nomor_unit,kondisi,qr_path,keterangan)
                    VALUES(%s,%s,%s,%s,%s,%s)""",
                    (barang_id,kode_unit,str(n),kondisi,qr,keterangan))
                added += 1
            # Aktifkan unit tracking di barang induk
            cur.execute("UPDATE barang SET has_unit_tracking=TRUE WHERE id=%s",(barang_id,))
            mysql.connection.commit(); cur.close()
            msg = f'{added} unit berhasil ditambahkan.'
            if skipped: msg += f' Dilewati (sudah ada): No. {", ".join(skipped)}.'
            flash(msg, 'success' if added else 'warning')
        else:
            # Tambah satu unit
            nomor     = request.form.get('nomor_unit','').strip()
            label     = request.form.get('label_unit','').strip()
            kondisi   = request.form.get('kondisi','Baik')
            keterangan= request.form.get('keterangan','').strip()
            kode_unit = generate_kode_unit(barang['kode_barang'], nomor)

            cur.execute("SELECT id FROM barang_unit WHERE kode_unit=%s",(kode_unit,))
            if cur.fetchone():
                flash(f'Unit No. {nomor} sudah ada.','danger')
                cur.close()
                return redirect(url_for('unit_tambah', barang_id=barang_id))

            qr = buat_qr_unit(kode_unit)
            cur.execute("""INSERT INTO barang_unit
                (barang_id,kode_unit,nomor_unit,label_unit,kondisi,qr_path,keterangan)
                VALUES(%s,%s,%s,%s,%s,%s,%s)""",
                (barang_id,kode_unit,nomor,label,kondisi,qr,keterangan))
            cur.execute("UPDATE barang SET has_unit_tracking=TRUE WHERE id=%s",(barang_id,))
            mysql.connection.commit(); cur.close()
            flash(f'Unit No. {nomor} ({kode_unit}) berhasil ditambahkan.','success')

        return redirect(url_for('unit_list', barang_id=barang_id))

    cur.close()
    return render_template('barang_unit_form.html', barang=barang, unit=None, mode='tambah')

# ============================================================
# UNIT BARANG — EDIT
# ============================================================
@app.route('/barang/<int:barang_id>/unit/<int:unit_id>/edit', methods=['GET','POST'])
@login_required
def unit_edit(barang_id, unit_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE id=%s",(barang_id,))
    barang = cur.fetchone()
    cur.execute("SELECT * FROM barang_unit WHERE id=%s AND barang_id=%s",(unit_id,barang_id))
    unit = cur.fetchone()
    if not barang or not unit:
        flash('Data tidak ditemukan.','danger')
        return redirect(url_for('barang_list'))

    if request.method == 'POST':
        label     = request.form.get('label_unit','').strip()
        kondisi   = request.form.get('kondisi','Baik')
        keterangan= request.form.get('keterangan','').strip()
        cur.execute("""UPDATE barang_unit
                       SET label_unit=%s, kondisi=%s, keterangan=%s
                       WHERE id=%s""",
                    (label, kondisi, keterangan, unit_id))
        mysql.connection.commit(); cur.close()
        flash(f'Unit No. {unit["nomor_unit"]} berhasil diperbarui.','success')
        return redirect(url_for('unit_list', barang_id=barang_id))

    cur.close()
    return render_template('barang_unit_form.html', barang=barang, unit=unit, mode='edit')

# ============================================================
# UNIT BARANG — HAPUS
# ============================================================
@app.route('/barang/<int:barang_id>/unit/<int:unit_id>/hapus', methods=['POST'])
@login_required
def unit_hapus(barang_id, unit_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang_unit WHERE id=%s AND barang_id=%s",(unit_id,barang_id))
    unit = cur.fetchone()
    if unit:
        # Hapus QR file
        if unit.get('qr_path'):
            f = os.path.join(app.static_folder, unit['qr_path'])
            if os.path.exists(f): os.remove(f)
        # Null-out referensi di peminjaman
        cur.execute("UPDATE peminjaman SET unit_id=NULL WHERE unit_id=%s",(unit_id,))
        cur.execute("DELETE FROM barang_unit WHERE id=%s",(unit_id,))
        # Jika semua unit dihapus, matikan has_unit_tracking
        cur.execute("SELECT COUNT(*) AS n FROM barang_unit WHERE barang_id=%s",(barang_id,))
        if cur.fetchone()['n'] == 0:
            cur.execute("UPDATE barang SET has_unit_tracking=FALSE WHERE id=%s",(barang_id,))
        mysql.connection.commit()
        flash(f'Unit No. {unit["nomor_unit"]} berhasil dihapus.','success')
    cur.close()
    return redirect(url_for('unit_list', barang_id=barang_id))

# ============================================================
# UNIT BARANG — DETAIL (target QR scan, publik)
# ============================================================
@app.route('/unit/detail/<kode_unit>')
def unit_detail(kode_unit):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT u.*, b.nama_barang, b.kategori, b.tipe,
                          b.spesifikasi, b.kode_barang AS kode_induk,
                          b.id AS barang_id_induk
                   FROM barang_unit u
                   JOIN barang b ON u.barang_id = b.id
                   WHERE u.kode_unit=%s""", (kode_unit,))
    unit = cur.fetchone()
    if not unit:
        return render_template('404.html'), 404

    # Peminjaman aktif untuk unit ini
    cur.execute("""SELECT p.id, p.kode_pinjam, p.nama_peminjam,
                          p.kelas_jabatan, p.tgl_pinjam,
                          p.tgl_kembali_rencana, p.status
                   FROM peminjaman p
                   WHERE p.unit_id=%s AND p.status IN ('Dipinjam','Terlambat')
                   ORDER BY p.tgl_pinjam DESC LIMIT 1""", (unit['id'],))
    peminjaman_aktif = cur.fetchone()

    # Riwayat peminjaman unit ini (10 terakhir)
    cur.execute("""SELECT p.kode_pinjam, p.nama_peminjam, p.kelas_jabatan,
                          p.tgl_pinjam, p.tgl_kembali_aktual, p.status, p.kondisi_kembali
                   FROM peminjaman p
                   WHERE p.unit_id=%s
                   ORDER BY p.tgl_pinjam DESC LIMIT 10""", (unit['id'],))
    riwayat = cur.fetchall()
    cur.close()

    return render_template('detail_unit.html',
        unit=unit, peminjaman_aktif=peminjaman_aktif, riwayat=riwayat)

# ============================================================
# UNIT BARANG — API (untuk scanner)
# ============================================================
@app.route('/api/unit/<kode_unit>')
def api_unit(kode_unit):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT u.*, b.nama_barang, b.kategori, b.tipe, b.kode_barang AS kode_induk
                   FROM barang_unit u JOIN barang b ON u.barang_id=b.id
                   WHERE u.kode_unit=%s""", (kode_unit,))
    u = cur.fetchone()
    if u:
        for k in ('created_at','updated_at'):
            if u.get(k): u[k] = str(u[k])
        # Cek peminjaman aktif
        cur.execute("""SELECT id, kode_pinjam, nama_peminjam, kelas_jabatan,
                              status, tgl_pinjam, tgl_kembali_rencana
                       FROM peminjaman WHERE unit_id=%s
                       AND status IN ('Dipinjam','Terlambat')
                       ORDER BY tgl_pinjam DESC LIMIT 1""", (u['id'],))
        p = cur.fetchone()
        cur.close()
        if p:
            for k in ('tgl_pinjam','tgl_kembali_rencana'):
                if p.get(k): p[k] = p[k].strftime('%d %b %Y, %H:%M')
            u['peminjaman_aktif'] = p
        else:
            u['peminjaman_aktif'] = None
        return jsonify({'status':'found','tipe':'unit','data':u})
    cur.close()
    return jsonify({'status':'not_found'}), 404

# ============================================================
# UNIT BARANG — CETAK LABEL (semua unit dalam satu halaman)
# ============================================================
@app.route('/barang/<int:barang_id>/unit/cetak-label')
@login_required
def unit_cetak_label(barang_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM barang WHERE id=%s",(barang_id,))
    barang = cur.fetchone()
    cur.execute("""SELECT * FROM barang_unit WHERE barang_id=%s
                   AND qr_path IS NOT NULL
                   ORDER BY CAST(nomor_unit AS UNSIGNED)""", (barang_id,))
    units = cur.fetchall(); cur.close()
    return render_template('unit_cetak_label.html', barang=barang, units=units)

# ============================================================
# ABOUT — Halaman Tentang Sistem
# ============================================================
@app.route('/about')
@login_required
def about():
    return render_template('about.html')

# ============================================================
# RIWAYAT PEMAKAIAN INVENTARIS TETAP
# Menampilkan gabungan: peminjaman + scan QR sebagai riwayat pemakaian
# ============================================================
@app.route('/barang/riwayat')
@login_required
def barang_riwayat():
    f_status = request.args.get('status', '').strip()
    f_period = request.args.get('period', '').strip()
    search   = request.args.get('search', '').strip()
    page     = int(request.args.get('page', 1))
    per_page = 15; offset = (page-1)*per_page

    date_from, period_label = get_date_range(f_period)

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append(
            "(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s "
            " OR COALESCE(b.nama_barang, p.kode_barang) LIKE %s)"
        )
        params += [like, like, like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    if date_from:
        clauses.append("p.tgl_pinjam >= %s"); params.append(date_from)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""
        SELECT
            p.id, p.kode_pinjam, p.kode_barang,
            COALESCE(b.nama_barang, p.kode_barang) AS nama_barang,
            COALESCE(b.kategori,    '(barang dihapus)') AS kategori,
            COALESCE(b.tipe,        '') AS tipe,
            p.nama_peminjam, p.kelas_jabatan, p.jumlah_pinjam,
            p.tgl_pinjam, p.tgl_kembali_rencana, p.tgl_kembali_aktual,
            p.status, p.kondisi_kembali, p.keperluan,
            CASE WHEN p.barang_id IS NULL THEN 1 ELSE 0 END AS barang_dihapus
        FROM peminjaman p
        LEFT JOIN barang b ON p.barang_id = b.id
        {where}
        ORDER BY p.tgl_pinjam DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    rows = cur.fetchall()

    cur.execute(f"""SELECT COUNT(*) AS n FROM peminjaman p
        LEFT JOIN barang b ON p.barang_id = b.id {where}
    """, params)
    total = cur.fetchone()['n']

    # Stats mengikuti filter periode yang aktif
    sp = [date_from] if date_from else []
    sw = "WHERE p.tgl_pinjam >= %s" if date_from else ""
    cur.execute(f"SELECT COUNT(*) AS n FROM peminjaman p {sw}", sp)
    s_total = cur.fetchone()['n']
    cur.execute(f"SELECT COUNT(*) AS n FROM peminjaman p {sw}" +
               (" AND " if sw else " WHERE ") + "p.status='Dipinjam'", sp)
    s_aktif = cur.fetchone()['n']
    cur.execute(f"SELECT COUNT(*) AS n FROM peminjaman p {sw}" +
               (" AND " if sw else " WHERE ") + "p.status='Dikembalikan'", sp)
    s_kembali = cur.fetchone()['n']
    cur.execute(f"SELECT COUNT(*) AS n FROM peminjaman p {sw}" +
               (" AND " if sw else " WHERE ") + "p.status='Terlambat'", sp)
    s_terlambat = cur.fetchone()['n']
    cur.close()

    return render_template('barang_riwayat.html',
        rows=rows, search=search, f_status=f_status,
        f_period=f_period, period_label=period_label,
        page=page, total_pages=max(1,(total+per_page-1)//per_page),
        total_rows=total,
        stats=dict(total=s_total, aktif=s_aktif,
                   kembali=s_kembali, terlambat=s_terlambat),
        PERIOD_LABELS=PERIOD_LABELS,
    )


# ============================================================
# RIWAYAT PEMAKAIAN INVENTARIS — EXPORT EXCEL
# ============================================================
@app.route('/barang/riwayat/export/excel')
@login_required
def barang_riwayat_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    f_status = request.args.get('status', '').strip()
    f_period = request.args.get('period', '').strip()

    date_from, period_label = get_date_range(f_period)
    search   = request.args.get('search', '').strip()

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append(
            "(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s "
            " OR COALESCE(b.nama_barang, p.kode_barang) LIKE %s)"
        )
        params += [like, like, like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    if date_from:
        clauses.append("p.tgl_pinjam >= %s"); params.append(date_from)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""
        SELECT
            p.*,
            COALESCE(b.nama_barang, p.kode_barang) AS nama_barang_display,
            COALESCE(b.tipe, '') AS tipe_display
        FROM peminjaman p
        LEFT JOIN barang b ON p.barang_id = b.id
        {where}
        ORDER BY p.created_at DESC
    """, params)
    rows = cur.fetchall(); cur.close()

    wb  = openpyxl.Workbook(); ws = wb.active
    ws.title = "Riwayat Pemakaian Inventaris"

    hfill  = PatternFill("solid", fgColor="1e3a8a")
    hfont  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='CBD5E1')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:M1')
    ws['A1'] = 'RIWAYAT PEMAKAIAN INVENTARIS — LAB TJKT SMKN 1 Cikarang Selatan'
    ws['A1'].font      = Font(name='Calibri', bold=True, size=14, color="1e3a8a")
    ws['A1'].alignment = center

    ws.merge_cells('A2:M2')
    label = f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}'
    if f_status: label += f'  |  Status: {f_status}'
    if f_period: label += f'  |  Periode: {period_label}'
    ws['A2'] = label
    ws['A2'].font      = Font(name='Calibri', size=10, italic=True, color="94a3b8")
    ws['A2'].alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    headers = ['No','Kode Pinjam','Nama Barang','Tipe','Nama Peminjam',
               'Kelas/Jabatan','Keperluan','Jml','Waktu Pinjam',
               'Rencana Kembali','Waktu Kembali','Kondisi Kembali','Status']
    widths  = [5, 20, 28, 16, 22, 14, 24, 6, 20, 20, 20, 14, 13]

    ws.append([]); ws.append(headers)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 22

    sfill = {'Dipinjam':'DBEAFE','Dikembalikan':'D1FAE5','Terlambat':'FEE2E2'}

    def fmt_dt(val):
        if not val: return '-'
        try: return val.strftime('%d/%m/%Y %H:%M')
        except: return str(val)

    for i, p in enumerate(rows, 1):
        ws.append([
            i, p['kode_pinjam'],
            p.get('nama_barang_display') or p['kode_barang'],
            p.get('tipe_display') or '-',
            p['nama_peminjam'],
            p['kelas_jabatan'] or '-',
            p['keperluan'] or '-',
            p['jumlah_pinjam'],
            fmt_dt(p['tgl_pinjam']),
            fmt_dt(p['tgl_kembali_rencana']),
            fmt_dt(p['tgl_kembali_aktual']),
            p['kondisi_kembali'] or '-',
            p['status'],
        ])
        rn = 4 + i; ws.row_dimensions[rn].height = 18
        for ci in range(1, 14):
            cell = ws.cell(row=rn, column=ci)
            cell.border = bdr
            cell.alignment = center if ci in (1, 8) else left
        sc = ws.cell(row=rn, column=13)
        fc = sfill.get(p['status'])
        if fc:
            sc.fill = PatternFill("solid", fgColor=fc)
            sc.font = Font(name='Calibri', bold=True, size=10)
            sc.alignment = center

    ws.freeze_panes = 'A5'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"RiwayatPemakaian_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================
# RIWAYAT PEMAKAIAN INVENTARIS — EXPORT PDF
# ============================================================
@app.route('/barang/riwayat/export/pdf')
@login_required
def barang_riwayat_export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    f_status = request.args.get('status', '').strip()
    search   = request.args.get('search', '').strip()

    clauses, params = [], []
    if search:
        like = f"%{search}%"
        clauses.append(
            "(p.nama_peminjam LIKE %s OR p.kode_pinjam LIKE %s "
            " OR COALESCE(b.nama_barang, p.kode_barang) LIKE %s)"
        )
        params += [like, like, like]
    if f_status:
        clauses.append("p.status=%s"); params.append(f_status)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    cur = mysql.connection.cursor()
    cur.execute(f"""
        SELECT p.*,
            COALESCE(b.nama_barang, p.kode_barang) AS nama_barang_display,
            COALESCE(b.tipe, '') AS tipe_display
        FROM peminjaman p
        LEFT JOIN barang b ON p.barang_id = b.id
        {where} ORDER BY p.created_at DESC
    """, params)
    rows = cur.fetchall(); cur.close()

    def fmt_dt(val):
        if not val: return '-'
        try: return val.strftime('%d/%m/%Y %H:%M')
        except: return str(val)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.8*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet(); elems = []
    navy = colors.HexColor('#1e3a8a')

    ts_h = ParagraphStyle('h', fontSize=13, textColor=navy, alignment=1,
                           fontName='Helvetica-Bold', spaceAfter=4)
    ts_s = ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#475569'),
                           alignment=1, spaceAfter=2)
    ts_d = ParagraphStyle('d', fontSize=7.5, textColor=colors.HexColor('#94a3b8'),
                           alignment=1, spaceAfter=10)

    judul = 'RIWAYAT PEMAKAIAN INVENTARIS — LAB TJKT'
    if f_status: judul += f' | Filter: {f_status}'
    elems += [
        Paragraph(judul, ts_h),
        Paragraph('SMKN 1 Cikarang Selatan, Bekasi', ts_s),
        Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}', ts_d),
        Spacer(1, 0.2*cm),
    ]

    header = ['No','Kode','Nama Barang','Peminjam','Kls/Jab',
              'Jml','Waktu Pinjam','Rencana Kembali','Kembali Aktual','Status']
    data   = [header]
    sfill_pdf = {
        'Dipinjam':     colors.HexColor('#DBEAFE'),
        'Dikembalikan': colors.HexColor('#D1FAE5'),
        'Terlambat':    colors.HexColor('#FEE2E2'),
    }
    ts_style = [
        ('BACKGROUND',    (0,0), (-1,0), navy),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE',      (0,1), (-1,-1), 7.5),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID',          (0,0), (-1,-1), 0.35, colors.HexColor('#CBD5E1')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ALIGN',         (2,1), (4,-1), 'LEFT'),
    ]

    for i, p in enumerate(rows, 1):
        data.append([
            str(i), p['kode_pinjam'],
            p.get('nama_barang_display') or p['kode_barang'],
            p['nama_peminjam'],
            p['kelas_jabatan'] or '-',
            str(p['jumlah_pinjam']),
            fmt_dt(p['tgl_pinjam']),
            fmt_dt(p['tgl_kembali_rencana']),
            fmt_dt(p['tgl_kembali_aktual']),
            p['status'],
        ])
        fc = sfill_pdf.get(p['status'])
        if fc:
            ts_style.append(('BACKGROUND', (9, i), (9, i), fc))
            ts_style.append(('FONTNAME',   (9, i), (9, i), 'Helvetica-Bold'))

    cw  = [0.7*cm,2.8*cm,4.5*cm,3.5*cm,2.2*cm,0.9*cm,2.8*cm,2.8*cm,2.8*cm,2.2*cm]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle(ts_style))
    elems.append(tbl)
    doc.build(elems); buf.seek(0)
    fname = f"RiwayatPemakaian_TJKT_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')

# ============================================================
# AUTO-REGENERATE QR — jalan otomatis saat request pertama
# Hanya memperbaiki QR lama yang masih pakai [IP_KOMPUTER]
# Tidak menyentuh QR yang sudah benar
# ============================================================
_qr_fixed = False   # flag agar hanya berjalan sekali per session

@app.before_request
def auto_fix_qr_on_first_request():
    global _qr_fixed
    if _qr_fixed:
        return                          # sudah berjalan, skip
    if request.endpoint in (None, 'static'):
        return                          # skip request aset statis
    _qr_fixed = True                    # tandai sudah berjalan

    try:
        cur = mysql.connection.cursor()

        # Perbaiki QR inventaris tetap yang masih pakai placeholder
        cur.execute("""SELECT id, kode_barang FROM barang
                       WHERE qr_path IS NULL
                          OR qr_path = ''""")
        for row in cur.fetchall():
            try:
                qp = buat_qr_code(row['kode_barang'])
                cur.execute("UPDATE barang SET qr_path=%s WHERE id=%s",
                            (qp, row['id']))
            except Exception:
                pass

        # Perbaiki QR habis pakai yang masih pakai placeholder
        cur.execute("""SELECT id, kode_barang FROM barang_habis_pakai
                       WHERE qr_path IS NULL
                          OR qr_path = ''""")
        for row in cur.fetchall():
            try:
                qp = buat_qr_consumable(row['kode_barang'])
                cur.execute("UPDATE barang_habis_pakai SET qr_path=%s WHERE id=%s",
                            (qp, row['id']))
            except Exception:
                pass

        mysql.connection.commit()
        cur.close()
    except Exception:
        pass    # gagal pun tidak menganggu jalannya aplikasi


if __name__=='__main__':
    app.run(host='0.0.0.0',port=5000,debug=True)